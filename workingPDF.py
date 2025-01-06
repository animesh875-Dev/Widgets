import json
import requests
from openpyxl import Workbook
import urllib3
from datetime import datetime
import os
import xml.etree.ElementTree as ET
import tkinter as tk 
from tkinter import ttk, messagebox
from cryptography.fernet import Fernet
from tkinter import simpledialog, messagebox  # Ensure that simpledialog is explicitly imported
from fpdf import FPDF

# Suppress warnings about unverified HTTPS requests (for testing purposes)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

with open("config.json", "r") as config_file:
    config = json.load(config_file)

server_url = config["server_url"]
data_governance_TP = config["data_governance_TP"]
data_governance_TC = config["data_governance_TC"] 
data_governance_TS = config["data_governance_TS"]
data_governance_TSuite = config["data_governance_TSuite"]
data_governance_TCER = config["data_governance_TCER"]

# Function to validate credentials (dummy API for validation in this example)
def validate_credentials(username, password):
  
    api_url = f"{server_url}/qm" 
    try:
        response = requests.get(api_url, auth=(username, password), verify=False)
        
        if response.status_code == 200:
            return True
        else:
            return False

    except requests.exceptions.RequestException as e:
        print(f"Error validating credentials: {e}")
        return False


# Function to ask user for credentials through a dialog box
def get_credentials():
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    username = simpledialog.askstring("Input", "Enter your Username:", parent=root)
    password = simpledialog.askstring("Input", "Enter your Password:", parent=root, show="*")
    return username, password


# Function to display an error message and terminate the program
def show_error_and_terminate(message):
    messagebox.showerror("Invalid Credentials", message)
    exit(0)


# Main script
username, password = get_credentials()

# Validate credentials
if not validate_credentials(username, password):
    show_error_and_terminate("You have entered an Invalid credentials. The session will now terminate.")

# Proceed with the rest of the code if credentials are valid
print(f"Credentials for {username} validated successfully.")


 # Assuming max allowed test cases is 100, replace with actual value if needed

# API endpoints
api_url_project_areas = f"{server_url}/qm/service/com.ibm.team.repository.service.internal.webuiInitializer.IWebUIInitializerRestService/initializationData"
oslc_api_url_template = f"{server_url}/qm/service/com.ibm.rqm.configmanagement.service.rest.IConfigurationManagementRestService/pagedSearchResult?pageSize=100&page=0&projectArea={{Project_Area_UUID}}"
api_url = f"{server_url}/qm/service/com.ibm.rqm.planning.common.service.rest.ITestCaseRestService/pagedSearchResult"
api_url_tcer = f"{server_url}/qm/service/com.ibm.rqm.execution.common.service.rest.ITestcaseExecutionRecordRestService/pagedSearchResult"

# File to save printed messages (initialized later dynamically)
MESSAGE_LOG_FILE = ""

def log_message_to_file(message, project_area_name):
    """
    Logs the provided message to a text file with project area and timestamp.
    """
    os.makedirs("Reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_name = f"Reports/{project_area_name}_{timestamp}.txt"
    with open(log_file_name, "a") as file:
        file.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

def fetch_project_areas():
    """
    Fetches project areas using the API.
    """
    try:
        print("Fetching project areas...")
        response = requests.get(api_url_project_areas, auth=(username, password), verify=False)
        response.raise_for_status()

        data = response.json()
        if 'soapenv:Body' in data and "response" in data['soapenv:Body']:
            return data['soapenv:Body']["response"]["returnValue"]["value"]["com.ibm.rqm.planning.service.permissionsWebUIInitializer"]["userProjectAreas"]
        else:
            print("Invalid response structure.")
            log_message_to_file("Invalid response structure while fetching project areas.")
    except requests.exceptions.RequestException as e:
        print(f"Error fetching project areas: {e}")
        log_message_to_file(f"Error fetching project areas: {e}")
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON response: {e}")
        log_message_to_file(f"Error decoding JSON response: {e}")

    return []

def fetch_oslc_details(project_area_uuid):
    """
    Fetches OSLC details for a given project area and retrieves all streams.
    """
    if not project_area_uuid:
        print("Error: Project Area UUID is missing.")
        log_message_to_file(f"Error: Project Area UUID is missing.", "Unknown")
        return []

    oslc_url = oslc_api_url_template.replace("{Project_Area_UUID}", project_area_uuid)
    try:
        print(f"Fetching OSLC details for project area {project_area_uuid}...")
        response = requests.get(oslc_url, auth=(username, password), verify=False)
        response.raise_for_status()

        root = ET.fromstring(response.text)
        result_set_size = int(root.find('.//resultSetSize').text)
        results = root.findall('.//results')

        streams = []
        for result in results[:result_set_size]:
            item_id = result.find("itemId").text if result.find("itemId") is not None else None
            name = result.find("name").text if result.find("name") is not None else None
            if item_id and name:
                streams.append({
                    "Project_Area_Stream_Name": name,
                    "Project_Area_Stream_OSLC_ID": item_id
                })

        return streams
    except ET.ParseError as e:
        print(f"Error parsing XML response for Project Area UUID {project_area_uuid}: {e}")
        log_message_to_file(f"Error parsing XML response for Project Area UUID {project_area_uuid}: {e}", project_area_uuid)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching OSLC details for Project Area UUID {project_area_uuid}: {e}")
        log_message_to_file(f"Error fetching OSLC details for Project Area UUID {project_area_uuid}: {e}", project_area_uuid)

    return []

def fetch_test_plan_count(project_area_uuid, oslc_id):
    """
    Fetches the test plan count for a given project area and OSLC ID.
    """
    test_plan_api_url = f"{server_url}/qm/service/com.ibm.rqm.planning.common.service.rest.ITestPlanRestService/pagedSearchResult?processArea={project_area_uuid}&page=0&pageSize=500&oslc_config.context={oslc_id}"
    try:
        print(f"Fetching test plan count for {project_area_uuid} and OSLC ID {oslc_id}...")
        response = requests.get(test_plan_api_url, auth=(username, password), verify=False)
        response.raise_for_status()

        root = ET.fromstring(response.text)
        result_set_size = int(root.find('.//resultSetSize').text)
        return result_set_size
    
    except ET.ParseError as e:
        print(f"Error parsing XML response for Test Plan Count: {e}")
        log_message_to_file(f"Error parsing XML response for Test Plan Count: {e}")
    except requests.exceptions.RequestException as e:
        print(f"Error fetching Test Plan Count: {e}")
        log_message_to_file(f"Error fetching Test Plan Count: {e}")

    return 0


# Headers
headers = {
    "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
    "Accept": "application/json",
}

def build_request_body(page=0, page_size=100, process_area="" ,oslc_context=""):
    """
    Build the body of the request dynamically based on parameters.
    """
    body = {
        "includeCustomAttributes": "true",
        "includeArchived": "false",
        "processArea": process_area,
        "traceabilityViewType": "true",
        "resolveParentTestPlans": "false",
        "resolveScripts": "false",
        "resolveParentTestSuites": "true",
        "resolveCategories": "true",
        "resolveCustomAttributes": "false",
        "resolveLinkedFiles": "false",
        "resolveDevItem": "false",
        "resolveCopiedArtifactInfo": "false",
        "page": str(page),
        "pageSize": str(page_size),
        "resultLimit": "-1",
        "oslc_config.context":oslc_context,
        "isWebUI": "true",
    }
    return "&".join(f"{key}={value}" for key, value in body.items())

def fetch_test_case_count(body):
    """
    Sends a POST request to the API and extracts the <totalSize> value from the response.
    """
    try:
        # Make the POST request
        response = requests.post(api_url, data=body, headers=headers, auth=(username, password), verify=False)
        response.raise_for_status()

        # Parse the XML response
        root = ET.fromstring(response.text)
        # body = build_request_body(page=1, page_size=50, process_area=Project_Area_UUID)        

        # Find the <totalSize> element
        total_size_element = root.find(".//totalSize")
        if total_size_element is not None:
            total_size = int(total_size_element.text)
            print(f"Total Test Case Count: {total_size}")
            return total_size
        else:
            print("No <totalSize> element found in the response.")
            return None
        print(f"Total Test case count: {total_size}")
    except requests.exceptions.RequestException as e:
        print(f"Error making API request: {e}")
        return None
    except ET.ParseError as e:
        print(f"Error parsing XML response: {e}")
        return None

def parse_project_areas(user_project_areas):
    """
    Parses the project areas into a simplified format.
    """
    return [
        {"Project_Area_Name": area.get("name"), "Project_Area_UUID": area.get("itemId")}
        for area in user_project_areas if area.get("name") and area.get("itemId")
    ]

def fetch_test_script_count(project_area_uuid, oslc_id):
    """
    Fetches the test script count for a given project area and OSLC ID.
    """
    test_script_api_url = f"{server_url}/qm/service/com.ibm.rqm.execution.common.service.rest.IExecutionScriptSearchRestService/pagedSearchResult"
    params = {
        'oslc_config.context': oslc_id,
        'pageSize': 500,
        'processArea': project_area_uuid,
    }
    try:
        print(f"Fetching test script count for {project_area_uuid} and OSLC ID {oslc_id}...")
        response = requests.get(test_script_api_url, params=params, auth=(username, password), verify=False)
        response.raise_for_status()

        root = ET.fromstring(response.text)
        total_size_element = root.find(".//totalSize")
        if total_size_element is not None:
            total_size = int(total_size_element.text)
            print(f"Total Test Script Count: {total_size}")
            return total_size
        else:
            print("No <totalSize> element found in the response.")
            return 0
    except requests.exceptions.RequestException as e:
        print(f"Error fetching Test Script Count: {e}")
        log_message_to_file(f"Error fetching Test Script Count: {e}")
    except ET.ParseError as e:
        print(f"Error parsing XML response for Test Script Count: {e}")
        log_message_to_file(f"Error parsing XML response for Test Script Count: {e}")

    return 0

def fetch_test_suite_count(project_area_id ,oslc_id):
    try:

        api_url_test_suite = f"{server_url}/qm/service/com.ibm.rqm.planning.common.service.rest.ITestSuiteRestService/pagedSearchResult"
        # Set the parameters for the API request
        params = {
            'oslc_config.context': oslc_id,
            'pageSize': 500,
            'processArea': project_area_id,
        }
        
        # Make the API request
        response = requests.get(api_url_test_suite, params=params, auth=(username, password), verify=False)
        response.raise_for_status()  # Raise an error for HTTP issues
        
        # Parse the XML response
        root = ET.fromstring(response.text)
        total_size_element = root.find(".//totalSize")
        test_suite_count = int(total_size_element.text) if total_size_element is not None else 0
        
        return test_suite_count
    except Exception as e:
        print(f"Error fetching test suite count: {e}")
        return None
# Headers
headers_tcer = {
    "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
    "Accept": "text/json",
}
def build_request_body_tcer(page=0, page_size=100, process_area="" ,oslc_context=""):
    """
    Build the body of the request dynamically based on parameters.
    """
    body = {
        "includeCustomAttributes": "true",
        "includeArchived": "false",
        "processArea": process_area,
        "traceabilityViewType": "true",
        "resolveParentTestPlans": "false",
        "resolveScripts": "false",
        "resolveParentTestSuites": "true",
        "resolveCategories": "true",
        "resolveCustomAttributes": "false",
        "resolveLinkedFiles": "false",
        "resolveDevItem": "false",
        "resolveCopiedArtifactInfo": "false",
        "page": str(page),
        "pageSize": str(page_size),
        "resultLimit": "-1",
        "oslc_config.context":oslc_context,
        "isWebUI": "true",
    }
    return "&".join(f"{key}={value}" for key, value in body.items())
def fetch_test_case_execution_record_count(body):

    try:
        response = requests.post(api_url_tcer, data=body, headers=headers_tcer, auth=(username, password), verify=False)
        response.raise_for_status()

        # Parse the JSON response
        
        data = response.json()
        if 'soapenv:Body' in data and "response" in data['soapenv:Body']:
            return data['soapenv:Body']["response"]["returnValue"]["value"]["totalSize"]
        else:
            print("Invalid response structure.")
            log_message_to_file("Invalid response structure while fetching project areas.")
    except requests.exceptions.RequestException as e:
        print(f"Error fetching project areas: {e}")
        log_message_to_file(f"Error fetching project areas: {e}")
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON response: {e}")
        log_message_to_file(f"Error decoding JSON response: {e}")

    return []

# Function to log messages to a file
def log_message_to_file(message, project_area_name):
    os.makedirs("Reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_name = f"Reports/{project_area_name}_{timestamp}.txt"
    with open(log_file_name, "a") as file:
        file.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

# Generate project report PDF
def generate_project_report(
    selected_project_area,
    selected_component,
    test_plan_count,
    data_governance_TP,
    test_case_count,
    data_governance_TC,
    test_script_count,
    data_governance_TS,
    test_suite_count,
    data_governance_TSuite,
    test_case_execution_record_count,
    data_governance_TCER
):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Title
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Project Summary Report", ln=True, align="C")
    pdf.ln(10)

    # Project Info
    project_info = [
        ("Project Area Name", selected_project_area),
        ("Stream", selected_component),
    ]
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(0, 10, "Project Information", ln=True)
    pdf.set_font("Arial", size=12)
    for label, value in project_info:
        pdf.cell(60, 10, f"{label}:", border=1)
        pdf.cell(0, 10, value, border=1, ln=True)

    pdf.ln(5)

    # Helper function to add a section
    def add_section(title, details):
        pdf.set_font("Arial", style="B", size=12)
        pdf.cell(0, 10, title, ln=True)
        pdf.set_font("Arial", size=12)
        for label, value in details:
            pdf.cell(100, 10, f"{label}:", border=1)
            pdf.cell(0, 10, value, border=1, ln=True)
        pdf.ln(5)

    # Test Plan Details
    if test_plan_count <= data_governance_TP:
        remaining_plans = data_governance_TP - test_plan_count
        test_plan_details = [
            ("Permission Status", "Allowed to create test plans"),
            ("Current Test Plan Count", str(test_plan_count)),
            ("Maximum Test Plans Allowed", str(data_governance_TP)),
            ("Remaining Test Plans", str(remaining_plans)),
        ]
    else:
        exceeded_plans = test_plan_count - data_governance_TP
        test_plan_details = [
            ("Permission Status", "Not allowed to create test plans"),
            ("Current Test Plan Count", str(test_plan_count)),
            ("Exceeded Test Plan Value", str(exceeded_plans)),
        ]
    add_section("Test Plan Details", test_plan_details)

    # Test Case Details
    if test_case_count <= data_governance_TC:
        remaining_cases = data_governance_TC - test_case_count
        test_case_details = [
            ("Permission Status", "Allowed to create test cases"),
            ("Current Test Case Count", str(test_case_count)),
            ("Maximum Test Cases Allowed", str(data_governance_TC)),
            ("Remaining Test Cases", str(remaining_cases)),
        ]
    else:
        exceeded_cases = test_case_count - data_governance_TC
        test_case_details = [
            ("Permission Status", "Not allowed to create test cases"),
            ("Current Test Case Count", str(test_case_count)),
            ("Exceeded Test Case Value", str(exceeded_cases)),
        ]
    add_section("Test Case Details", test_case_details)

    # Test Script Details
    if test_script_count <= data_governance_TS:
        remaining_scripts = data_governance_TS - test_script_count
        test_script_details = [
            ("Permission Status", "Allowed to create test scripts"),
            ("Current Test Script Count", str(test_script_count)),
            ("Maximum Test Scripts Allowed", str(data_governance_TS)),
            ("Remaining Test Scripts", str(remaining_scripts)),
        ]
    else:
        exceeded_scripts = test_script_count - data_governance_TS
        test_script_details = [
            ("Permission Status", "Not allowed to create test scripts"),
            ("Current Test Script Count", str(test_script_count)),
            ("Exceeded Test Script Value", str(exceeded_scripts)),
        ]
    add_section("Test Script Details", test_script_details)

    # Test Suite Details
    if test_suite_count <= data_governance_TSuite:
        remaining_suites = data_governance_TSuite - test_suite_count
        test_suite_details = [
            ("Permission Status", "Allowed to create test suites"),
            ("Current Test Suite Count", str(test_suite_count)),
            ("Maximum Test Suites Allowed", str(data_governance_TSuite)),
            ("Remaining Test Suites", str(remaining_suites)),
        ]
    else:
        exceeded_suites = test_suite_count - data_governance_TSuite
        test_suite_details = [
            ("Permission Status", "Not allowed to create test suites"),
            ("Current Test Suite Count", str(test_suite_count)),
            ("Exceeded Test Suite Value", str(exceeded_suites)),
        ]
    add_section("Test Suite Details", test_suite_details)

    # Test Case Execution Record Details
    if test_case_execution_record_count <= data_governance_TCER:
        remaining_records = data_governance_TCER - test_case_execution_record_count
        test_case_execution_details = [
            ("Permission Status", "Allowed to create test case execution records"),
            ("Current Test Case Execution Record Count", str(test_case_execution_record_count)),
            ("Maximum Test Case Execution Records Allowed", str(data_governance_TCER)),
            ("Remaining Test Case Execution Records", str(remaining_records)),
        ]
    else:
        exceeded_records = test_case_execution_record_count - data_governance_TCER
        test_case_execution_details = [
            ("Permission Status", "Not allowed to create test case execution records"),
            ("Current Test Case Execution Record Count", str(test_case_execution_record_count)),
            ("Exceeded Test Case Execution Record Value", str(exceeded_records)),
        ]
    add_section("Test Case Execution Record Details", test_case_execution_details)

    # Combine messages for both Test Plan and Test Case
    full_message = f"{test_plan_details}\n\n{test_case_details}\n\n{test_script_details}\n\n{test_suite_details}\n\n{test_case_execution_details}"

    # Display the message in a popup
    messagebox.showinfo("Test Plan and Test Case Validation", full_message)

    # Log the message to a file
    log_message_to_file(full_message, selected_project_area)
        

    # Combine messages for both Test Plan and Test Case
        

    # Save the PDF
    pdf_output_filename = f"Reports/Project_Summary_Report_{selected_project_area}_{selected_component}.pdf"
    pdf.output(pdf_output_filename)
    print(f"Report generated: {pdf_output_filename}")


def on_validate_data_click():
    """
    Handles the logic when the 'Validate Data' button is clicked.
    """
    selected_project_area = project_area_combobox.get()
    if selected_project_area:
        project_area_uuid = next(area["Project_Area_UUID"] for area in project_areas if area["Project_Area_Name"] == selected_project_area)
        selected_component = component_combobox.get()
        selected_oslc_id = next(comp["Project_Area_Stream_OSLC_ID"] for comp in components if comp["Project_Area_Stream_Name"] == selected_component)
        body = build_request_body(page=1, page_size=50, process_area=project_area_uuid ,oslc_context=selected_oslc_id)
        body_tcer = build_request_body_tcer(page=1, page_size=50, process_area=project_area_uuid ,oslc_context=selected_oslc_id)
        total_size = fetch_test_case_count(body)
        # Fetch counts
        test_plan_count = fetch_test_plan_count(project_area_uuid,selected_oslc_id)
        test_case_count = fetch_test_case_count(body)
        test_script_count = fetch_test_script_count(project_area_uuid,selected_oslc_id)
        test_suite_count = fetch_test_suite_count(project_area_uuid,selected_oslc_id)
        test_case_execution_record_count = fetch_test_case_execution_record_count(body_tcer)
        print(f"Total Count of test plan {test_plan_count}")
        print(f"Total Count of test case {test_case_count}")        
        print(f"Total Count of test script {test_script_count}") 
        print(f"Total Count of test suite {test_suite_count}") 
        print(f"Total Count of test case execution record {test_case_execution_record_count}") 



        # Generate the report PDF
        generate_project_report(selected_project_area, selected_component, test_plan_count, data_governance_TP, test_case_count, data_governance_TC,
                                test_script_count, data_governance_TS, test_suite_count, data_governance_TSuite, test_case_execution_record_count, data_governance_TCER)
    

        

def on_project_area_select(event):
    """
    Fetch and update components based on selected project area.
    """
    selected_project_area = project_area_combobox.get()
    if selected_project_area:
        project_area_uuid = next(area["Project_Area_UUID"] for area in project_areas if area["Project_Area_Name"] == selected_project_area)
        global components
        components = fetch_oslc_details(project_area_uuid)

        # Update the components dropdown
        component_combobox['values'] = [comp["Project_Area_Stream_Name"] for comp in components]
        component_combobox.set('')  # Clear previous selection
        component_combobox.grid(row=2, column=1, padx=10, pady=10)
        component_combobox.bind("<<ComboboxSelected>>")

        # Display the selected project area name
        selected_project_area_label.config(text=f"Selected Project Area: {selected_project_area}")

# Set up Tkinter window
window = tk.Tk()
window.title("Project Area and Stream Selector")
window.geometry("800x600")

# Fetch project areas and parse them
user_project_areas = fetch_project_areas()
if user_project_areas:
    project_areas = parse_project_areas(user_project_areas)

    # Project Area selection dropdown
    project_area_combobox = ttk.Combobox(window, values=[area["Project_Area_Name"] for area in project_areas], state="readonly", width=40)
    project_area_combobox.grid(row=0, column=1, padx=10, pady=10)
    project_area_combobox.bind("<<ComboboxSelected>>", on_project_area_select)

    # Label for project area selection
    project_area_label = tk.Label(window, text="Select Project Area:")
    project_area_label.grid(row=0, column=0, padx=10, pady=10)

    # Label for selected project area
    selected_project_area_label = tk.Label(window, text="Selected Project Area: ")
    selected_project_area_label.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

    # Components dropdown
    component_combobox = ttk.Combobox(window, state="readonly", width=40)
    component_combobox.grid(row=2, column=1, padx=10, pady=10)

    # Label for components selection
    component_label = tk.Label(window, text="Select Components:")
    component_label.grid(row=2, column=0, padx=10, pady=10)


    # Add a button to validate the data
    validate_button = tk.Button(window, text="Validate Data", command=on_validate_data_click)
    validate_button.grid(row=3, column=1, padx=10, pady=10)

    # Start Tkinter main loop
    window.mainloop()

else:
    messagebox.showerror("Error", "Failed to fetch project areas.")